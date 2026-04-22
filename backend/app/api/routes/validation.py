"""
API Routes - Validation workflow using LangGraph.
Candidates are auto-created from onboarding form.
"""
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from pydantic import BaseModel
from typing import Dict, Any, Optional, List
import os
import re
import csv
import io
import json
import aiofiles
import pypdf
import fitz
from datetime import datetime
from cryptography.fernet import Fernet
from fastapi.responses import StreamingResponse

from sqlalchemy.orm import Session
from app.core.database import get_db, Candidate
from app.core.config import config
from app.langgraph.orchestration import run_extraction_workflow, run_validation_workflow
import pypdf

router = APIRouter()

UPLOAD_DIR = config.UPLOAD_DIR
os.makedirs(UPLOAD_DIR, exist_ok=True)


def generate_candidate_id(name: str, email: str = "") -> str:
    """Generate unique ID from candidate name + email (for deduplication)."""
    clean_name = re.sub(r'[^a-z0-9]', '_', name.lower().strip())
    clean_name = re.sub(r'_+', '_', clean_name).strip('_')
    
    # Use email hash for uniqueness (same person = same ID)
    if email:
        import hashlib
        email_hash = hashlib.md5(email.lower().strip().encode()).hexdigest()[:6]
        return f"{clean_name}_{email_hash}"
    
    # Fallback: use timestamp
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return f"{clean_name}_{timestamp}"


# ============ ONBOARDING FORM UPLOAD ============

@router.post("/onboarding/upload")
async def upload_onboarding_form(
    form_file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Upload onboarding form (CSV/Excel) with multiple candidates.
    Auto-creates candidates from the form.
    """
    filename = (form_file.filename or "").lower()
    content = await form_file.read()
    
    candidates_created = []
    
    if filename.endswith('.csv'):
        decoded = content.decode('utf-8')
        reader = csv.DictReader(io.StringIO(decoded))
        
        for row in reader:
            # Normalize keys using LangGraph validation tools
            from app.langgraph.subgraphs.validation.tools import normalize_form_data
            raw_form = {k: v.strip() for k, v in row.items() if k and v}
            form_data = normalize_form_data(raw_form)
            
            # Get name and email using canonical field names
            name = form_data.get('full_name') or form_data.get('name') or form_data.get('candidate_name')
            email = form_data.get('email', '')
            if not name:
                continue
            
            candidate_id = generate_candidate_id(name, email)
            
            # Check if exists (by ID or email)
            existing = db.query(Candidate).filter(
                (Candidate.id == candidate_id) | (Candidate.email == email)
            ).first() if email else db.query(Candidate).filter(Candidate.id == candidate_id).first()
            
            if existing:
                # Merge form data with existing
                old_form = existing.get_form() or {}
                for k, v in form_data.items():
                    if v and (k not in old_form or not old_form[k]):
                        old_form[k] = v
                    elif v:
                        old_form[k] = v  # Override with new value
                existing.set_form(old_form)
                candidates_created.append({"id": candidate_id, "name": name, "status": "updated"})
            else:
                # Split name into first/last for the new unified model
                parts = name.split(maxsplit=1)
                first = parts[0] if parts else name
                last = parts[1] if len(parts) > 1 else ""

                candidate = Candidate(
                    id=candidate_id,
                    first_name=first,
                    last_name=last,
                    full_name=name,
                    email=form_data.get('email', ''),
                    phone=form_data.get('phone', '')
                )
                candidate.set_form(form_data)
                db.add(candidate)
                candidates_created.append({"id": candidate_id, "name": name, "status": "created"})
        
        db.commit()
    
    elif filename.endswith(('.xlsx', '.xls')):
        try:
            import openpyxl
            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(content))
            ws = wb.active
            
            headers = [str(c.value).lower().replace(' ', '_') if c.value else '' for c in ws[1]]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                form_data = {headers[i]: str(v) for i, v in enumerate(row) if i < len(headers) and v}
                
                name = form_data.get('full_name') or form_data.get('name')
                if not name:
                    continue
                
                candidate_id = generate_candidate_id(name)
                
                existing = db.query(Candidate).filter(Candidate.id == candidate_id).first()
                if existing:
                    existing.set_form(form_data)
                    candidates_created.append({"id": candidate_id, "name": name, "status": "updated"})
                else:
                    # Split name into first/last
                    parts = name.split(maxsplit=1)
                    first = parts[0] if parts else name
                    last = parts[1] if len(parts) > 1 else ""

                    candidate = Candidate(
                        id=candidate_id,
                        first_name=first,
                        last_name=last,
                        full_name=name,
                        email=form_data.get('email', ''),
                        phone=form_data.get('phone', '')
                    )
                    candidate.set_form(form_data)
                    db.add(candidate)
                    candidates_created.append({"id": candidate_id, "name": name, "status": "created"})
            
            db.commit()
        except ImportError:
            raise HTTPException(status_code=400, detail="Use CSV format (Excel requires openpyxl)")
    else:
        raise HTTPException(status_code=400, detail="Upload CSV or Excel file")
    
    return {
        "status": "success",
        "candidates_count": len(candidates_created),
        "candidates": candidates_created
    }


# ============ DOCUMENT UPLOAD ============

@router.post("/documents/{candidate_id}")
async def upload_documents(
    candidate_id: str,
    resume: Optional[UploadFile] = File(None),
    hr_transcript: Optional[UploadFile] = File(None),
    hr_audio: Optional[UploadFile] = File(None),
    aadhar: Optional[UploadFile] = File(None),
    pan: Optional[UploadFile] = File(None),
    marksheet_10th: Optional[UploadFile] = File(None),
    marksheet_12th: Optional[UploadFile] = File(None),
    i9_form: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    """Upload documents for a candidate."""
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found. Upload form first.")
    
    candidate_dir = os.path.join(UPLOAD_DIR, candidate_id)
    os.makedirs(candidate_dir, exist_ok=True)
    
    docs = candidate.get_documents()
    uploaded = []
    
    files = [
        ("resume", resume),
        ("hr_transcript", hr_transcript),
        ("hr_audio", hr_audio),
        ("aadhar", aadhar),
        ("pan", pan),
        ("marksheet_10th", marksheet_10th),
        ("marksheet_12th", marksheet_12th),
        ("i9_form", i9_form)
    ]
    
    for name, file in files:
        if file and file.filename:
            ext = os.path.splitext(file.filename)[1] or '.txt'
            path = os.path.join(candidate_dir, f"{name}{ext}")
            content = await file.read()
            
            # Phase 2: Document Forensics (Check for metadata tampering in PDFs before encryption)
            if ext.lower() == '.pdf':
                try:
                    pdf_reader = pypdf.PdfReader(io.BytesIO(content))
                    meta = pdf_reader.metadata
                    reader = pypdf.PdfReader(io.BytesIO(content))
                    meta = reader.metadata or {}
                    creator = str(meta.get("/Creator", "")).lower()
                    producer = str(meta.get("/Producer", "")).lower()
                    
                    if "photoshop" in creator or "illustrator" in creator or "photoshop" in producer:
                        alert = f"TAMPER RISK [High]: {name.upper()} was modified using graphic design software ({producer or creator})."
                        if alert not in docs.get("forensic_alerts", []):
                            docs["forensic_alerts"] = docs.get("forensic_alerts", []) + [alert]
                    elif "gimp" in creator or "gimp" in producer or "canva" in producer:
                        alert = f"TAMPER RISK [Medium]: {name.upper()} was exported from consumer design software ({producer or creator})."
                        if alert not in docs.get("forensic_alerts", []):
                            docs["forensic_alerts"] = docs.get("forensic_alerts", []) + [alert]
                except Exception:
                    pass
            # ---------------------------------------
            
            # Encrypt the file content before saving
            fernet = Fernet(config.FERNET_KEY.encode('utf-8'))
            encrypted_data = fernet.encrypt(content)
            
            async with aiofiles.open(path, "wb") as f:
                await f.write(encrypted_data)
            
            docs[name] = path
            uploaded.append(name)
    
    candidate.set_documents(docs)
    db.commit()
    
    return {
        "status": "success",
        "candidate_id": candidate_id,
        "uploaded": uploaded,
        "tamper_warning": candidate.tamper_warning if hasattr(candidate, 'tamper_warning') else False
    }

@router.delete("/documents/{candidate_id}")
async def reset_documents(candidate_id: str, db: Session = Depends(get_db)):
    """Reset all uploaded documents, knowledge base, and validation data for a candidate."""
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found.")
    
    candidate_dir = os.path.join(UPLOAD_DIR, candidate_id)
    if os.path.exists(candidate_dir):
        import shutil
        shutil.rmtree(candidate_dir)
        
    candidate.set_documents({})
    candidate.set_knowledge_base({})
    candidate.set_validation({})
    candidate.validation_score = None
    candidate.is_validated = False
    
    db.commit()
    
    return {"status": "success", "message": "Candidate documents and validation data reset."}


# ============ EXTRACTION (via LangGraph) ============

@router.post("/extract/{candidate_id}")
async def extract_knowledge_base(
    candidate_id: str,
    db: Session = Depends(get_db)
):
    """Extract knowledge base using the Hetro Extraction Engine and bridge it to SQLite."""
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
    
    docs = candidate.get_documents()
    if not docs:
        raise HTTPException(status_code=400, detail="No documents uploaded")
        
    async def event_generator():
        import tempfile
        import asyncio
        from cryptography.fernet import Fernet
        from candidate.services.document_service import DocumentService
        from candidate.db.database import SessionLocal as PgSessionLocal
        
        # Import Vector & NoSQL Storage
        from embeddings.embedding_generator import generate_embeddings
        from storage.chroma_handler import chroma_handler
        from storage.mongo_handler import mongo_handler
        
        yield "data: {\"type\": \"log\", \"message\": \"> Initializing Hetro Extraction Engine...\"}\n\n"
        
        try:
            knowledge_base = dict(candidate.get_knowledge_base())
            pg_db = PgSessionLocal()
            fernet = Fernet(config.FERNET_KEY.encode('utf-8'))
            
            yield "data: {\"type\": \"log\", \"message\": \"> Bridging SQLite and PostgreSQL Databases...\"}\n\n"
            
            for doc_name, path in docs.items():
                if doc_name == "forensic_alerts" or not isinstance(path, str):
                    continue
                    
                if not os.path.exists(path):
                    continue
                
                yield f"data: {{\"type\": \"log\", \"message\": \"> Decrypting {doc_name.upper()} into volatile memory...\"}}\n\n"
                
                # Decrypt the document
                async with aiofiles.open(path, "rb") as f:
                    encrypted_data = await f.read()
                
                decrypted_data = fernet.decrypt(encrypted_data)
                
                # Get the extension
                ext = os.path.splitext(path)[1] or '.pdf'
                
                # Create a temporary file to feed to DocumentService
                with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as temp_file:
                    temp_file.write(decrypted_data)
                    temp_path = temp_file.name
                
                try:
                    yield f"data: {{\"type\": \"stream\", \"message\": \"Extracting text and JSON structures from {doc_name.upper()} via Hetro Engine...\"}}\n\n"
                    
                    # Process the document with Hetro Engine
                    extraction_result = await DocumentService.extract_and_store_document(
                        db=pg_db,
                        file_path=temp_path,
                        candidate_id=candidate_id
                    )
                    
                    print(f"DEBUG: Extraction Result for {doc_name}: {extraction_result.extracted_data}")
                    
                    # Append extracted data to the knowledge base dictionary
                    knowledge_base[doc_name] = extraction_result.extracted_data
                    yield f"data: {{\"type\": \"log\", \"message\": \"> Successfully parsed {doc_name.upper()} into Knowledge Base.\"}}\n\n"
                    
                    # Vectorize and Store
                    raw_text = extraction_result.raw_text
                    if raw_text and len(raw_text.strip()) > 50:
                        yield f"data: {{\"type\": \"log\", \"message\": \"> Vectorizing {doc_name.upper()} text...\"}}\n\n"
                        try:
                            # 1. Generate Embeddings & Store in ChromaDB
                            embed_id = f"{candidate_id}_{doc_name}"
                            embeddings = generate_embeddings([raw_text])
                            if embeddings and len(embeddings) > 0 and len(embeddings[0]) > 0:
                                chroma_handler.store_embeddings(
                                    ids=[embed_id],
                                    embeddings=embeddings,
                                    documents=[raw_text],
                                    metadatas=[{"candidate_id": candidate_id, "doc_type": doc_name}]
                                )
                            
                            # 2. Store Raw Data in MongoDB
                            mongo_handler.store_document_metadata({
                                "candidate_id": candidate_id,
                                "doc_type": doc_name,
                                "raw_text": raw_text,
                                "parsed_json": extraction_result.extracted_data
                            })
                            yield f"data: {{\"type\": \"log\", \"message\": \"> Vector generated and indexed in ChromaDB & MongoDB.\"}}\n\n"
                        except Exception as storage_err:
                            yield f"data: {{\"type\": \"log\", \"message\": \"> ⚠ Vector storage warning for {doc_name.upper()}: {str(storage_err)}\"}}\n\n"
                
                except Exception as ex:
                    yield f"data: {{\"type\": \"log\", \"message\": \"> ⚠ Error extracting {doc_name.upper()}: {str(ex)}\"}}\n\n"
                finally:
                    # Clean up the temporary file immediately
                    if os.path.exists(temp_path):
                        os.remove(temp_path)
            
            pg_db.close()
            
            yield "data: {\"type\": \"log\", \"message\": \"> Finalizing parsed data schemas...\"}\n\n"
            
            # Save the newly constructed knowledge base back to SQLite Candidate
            from app.core.database import SessionLocal
            sqlite_db = SessionLocal()
            try:
                fresh_cand = sqlite_db.query(Candidate).filter(Candidate.id == candidate_id).first()
                if fresh_cand:
                    fresh_cand.set_knowledge_base(knowledge_base)
                    sqlite_db.commit()
                    yield "data: {\"type\": \"log\", \"message\": \"> Knowledge Base successfully committed to OnboardGuard Database.\"}\n\n"
            except Exception as dbe:
                yield f"data: {{\"type\": \"log\", \"message\": \"> DB Error: {str(dbe)}\"}}\n\n"
            finally:
                sqlite_db.close()
            
            sources = list(knowledge_base.keys())
            yield "data: {\"type\": \"log\", \"message\": \"> Extraction Engine Shutdown. Complete.\"}\n\n"
            
            final_resp = {
                "type": "result",
                "sources_extracted": sources,
                "knowledge_base": knowledge_base
            }
            yield f"data: {json.dumps(final_resp)}\n\n"
            
        except Exception as e:
            yield f"data: {{\"type\": \"log\", \"message\": \"> Critical Error: {str(e)}\"}}\n\n"
            yield "data: {\"type\": \"error\", \"message\": \"Extraction failed\"}\n\n"
    
    return StreamingResponse(event_generator(), media_type="text/event-stream")



# ============ PII REDACTION ============

@router.get("/documents/{candidate_id}/{doc_name}/redacted")
async def get_redacted_document(
    candidate_id: str,
    doc_name: str,
    db: Session = Depends(get_db)
):
    """Zero-Trust PII Auto-Redaction: serve redacted PDFs on the fly."""
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
    
    docs = candidate.get_documents()
    if doc_name not in docs:
        raise HTTPException(status_code=404, detail="Document not found")
        
    path = docs[doc_name]
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File missing on disk")
        
    async with aiofiles.open(path, "rb") as f:
        encrypted_data = await f.read()
        
    fernet = Fernet(config.FERNET_KEY.encode('utf-8'))
    try:
        decrypted_data = fernet.decrypt(encrypted_data)
    except Exception:
        raise HTTPException(status_code=500, detail="Failed to decrypt document")
        
    # Only process PDF files
    if not path.lower().endswith('.pdf'):
        import mimetypes
        from fastapi.responses import Response
        
        # Deduce true MIME type (e.g. text/plain for .txt, image/png for .png)
        mime_type, _ = mimetypes.guess_type(path)
        if not mime_type:
            mime_type = "application/octet-stream"
            
        return Response(
            content=decrypted_data, 
            media_type=mime_type,
            headers={"Content-Disposition": f'inline; filename="{doc_name}"'}
        )
        
    # Phase 3: Redact using fitz (PyMuPDF)
    doc = fitz.open(stream=decrypted_data, filetype="pdf")
    
    # Regex for 12-digit Aadhar and 10-char PAN
    aadhar_pattern = re.compile(r'\b\d{4}\s?\d{4}\s?\d{4}\b')
    pan_pattern = re.compile(r'\b[A-Z]{5}[0-9]{4}[A-Z]{1}\b')
    
    for page in doc:
        text = page.get_text("text")
        
        matches = []
        for match in aadhar_pattern.finditer(text):
            matches.append(match.group(0))
        for match in pan_pattern.finditer(text):
            matches.append(match.group(0))
            
        for match_str in matches:
            # Find bounds for each matched string
            areas = page.search_for(match_str)
            for area in areas:
                page.add_redact_annot(area, fill=(1, 1, 1))
        
        page.apply_redactions()
        
    redacted_pdf_bytes = doc.write()
    from fastapi.responses import Response
    return Response(
        content=redacted_pdf_bytes, 
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="redacted_{doc_name}.pdf"'}
    )


# ============ VALIDATION (via LangGraph) ============

@router.post("/validate/{candidate_id}")
async def validate_candidate(
    candidate_id: str,
    db: Session = Depends(get_db)
):
    """Validate candidate's onboarding form against knowledge base using LangGraph."""
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
    
    kb = candidate.get_knowledge_base()
    form = candidate.get_form()
    
    if not kb:
        raise HTTPException(status_code=400, detail="Extract knowledge base first")
    if not form:
        raise HTTPException(status_code=400, detail="No onboarding form data")
    
    # Fetch existing validation to preserve manual overrides
    existing_validation_data = candidate.get_validation()
    existing_validations_list = existing_validation_data.get("validations", []) if existing_validation_data else []
    
    # Run validation through LangGraph
    result = await run_validation_workflow(kb, form, existing_validations_list)
    
    # Store validation result
    validation_data = {
        "validations": result.get("validations", []),
        "overall_score": result.get("overall_score", 0),
        "correct_count": result.get("correct_count", 0),
        "incorrect_count": result.get("incorrect_count", 0),
        "ambiguous_count": result.get("ambiguous_count", 0)
    }
    candidate.set_validation(validation_data)
    candidate.validation_score = validation_data["overall_score"]
    db.commit()
    
    return {
        "status": "success",
        "candidate_id": candidate_id,
        "candidate_name": candidate.full_name,
        "summary": {
            "overall_score": validation_data["overall_score"],
            "correct": validation_data["correct_count"],
            "incorrect": validation_data["incorrect_count"],
            "ambiguous": validation_data["ambiguous_count"]
        },
        "details": validation_data
    }


# ============ RESOLVE AMBIGUOUS FIELDS ============

@router.post("/resolve/{candidate_id}")
async def resolve_ambiguous(
    candidate_id: str,
    body: dict,
    db: Session = Depends(get_db)
):
    """Mark an AMBIGUOUS field as CORRECT or INCORRECT."""
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
    
    field = body.get("field")
    resolution = body.get("resolution", "").upper()
    
    if resolution not in ("CORRECT", "INCORRECT"):
        raise HTTPException(status_code=400, detail="Resolution must be CORRECT or INCORRECT")
    
    validation_data = candidate.get_validation()
    if not validation_data:
        raise HTTPException(status_code=400, detail="No validation result to resolve")
    
    # Find and update the field
    validations = validation_data.get("validations", [])
    updated = False
    for v in validations:
        if v.get("field") == field and v.get("status") == "AMBIGUOUS":
            v["status"] = resolution
            v["reason"] = f"Manually marked as {resolution} by reviewer"
            updated = True
            break
    
    if not updated:
        raise HTTPException(status_code=404, detail="Ambiguous field not found")
    
    # Recalculate counts
    correct = sum(1 for v in validations if v["status"] == "CORRECT")
    incorrect = sum(1 for v in validations if v["status"] == "INCORRECT")
    ambiguous = sum(1 for v in validations if v["status"] == "AMBIGUOUS")
    total = correct + incorrect + ambiguous
    score = round((correct / total * 100), 1) if total else 0
    
    validation_data["validations"] = validations
    validation_data["correct_count"] = correct
    validation_data["incorrect_count"] = incorrect
    validation_data["ambiguous_count"] = ambiguous
    validation_data["overall_score"] = score
    
    candidate.set_validation(validation_data)
    candidate.validation_score = score
    db.commit()
    
    return {"status": "success", "field": field, "resolution": resolution, "new_score": score}


# ============ CANDIDATE MANAGEMENT ============

@router.get("/candidates")
async def list_candidates(
    q: str = "",
    db: Session = Depends(get_db)
):
    """List/search candidates. Use ?q=query to search by name/email."""
    query = db.query(Candidate)
    
    if q:
        search = f"%{q}%"
        query = query.filter(
            (Candidate.full_name.ilike(search)) | 
            (Candidate.email.ilike(search)) |
            (Candidate.id.ilike(search))
        )
    
    candidates = query.order_by(Candidate.created_at.desc()).limit(50).all()
    return {
        "candidates": [c.to_dict() for c in candidates],
        "total": len(candidates)
    }


@router.get("/candidate/{candidate_id}")
async def get_candidate(candidate_id: str, db: Session = Depends(get_db)):
    """Get full candidate details."""
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
    
    return {
        **candidate.to_dict(),
        "knowledge_base": candidate.get_knowledge_base(),
        "form_data": candidate.get_form(),
        "validation_result": candidate.get_validation() if getattr(candidate, 'is_validated', False) else None
    }

@router.get("/candidate/by-email/{email}")
async def get_candidate_by_email(email: str, db: Session = Depends(get_db)):
    """Get candidate details by email (bridge for ATS portal)."""
    candidate = db.query(Candidate).filter(Candidate.email == email).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found in OnboardGuard")
    
    return {
        **candidate.to_dict(),
        "knowledge_base": candidate.get_knowledge_base(),
        "form_data": candidate.get_form(),
        "validation_result": candidate.get_validation() if getattr(candidate, 'is_validated', False) else None
    }


@router.delete("/candidate/{candidate_id}")
async def delete_candidate(candidate_id: str, db: Session = Depends(get_db)):
    """Delete a candidate."""
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
    
    db.delete(candidate)
    db.commit()
    
    return {"status": "deleted", "candidate_id": candidate_id}
